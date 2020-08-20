import pytest
import openpyxl
import requests
import json
from jsonpath_ng import jsonpath, parse

'''
This routine loads the data from excel file as List of Dictionary. 
'''


def get_data(test_type):

    if test_type == 'smoke':
        data_sheet = "Smoke"
    elif test_type == 'regression':
        data_sheet = "Regression"
    else:
        data_sheet = "Functional"

    book = openpyxl.load_workbook("Data.xlsx")
    sheet = book[data_sheet]
    testcases=[]
    for row in range(2,sheet.max_row+1):
        test={}
        for column in range(1, sheet.max_column+1):
            test[sheet.cell(row=1,column=column).value]=sheet.cell(row=row,column=column).value  # adding as dict
        testcases.append(test)
    return testcases


'''
Below is the test that runs 'n' number of times based on the number of lines on the spreadsheet.
'''


@pytest.mark.smoke
@pytest.mark.parametrize('tests',get_data("smoke"))
def test_return_code(tests):

    response = requests.get(tests['url'])
    assert response.status_code == tests['return_code']


@pytest.mark.regression
@pytest.mark.parametrize('tests',get_data("regression"))
def test_regression(tests):
    response = requests.get(tests['url'])
    response_body = json.loads(response.text)
    assert response_body == json.loads(tests['expected'])      # comparing the API response vs expected JSON from excel.


@pytest.mark.functional
@pytest.mark.parametrize('tests', get_data("functional"))
def test_functional(tests):
    global response
    if tests['url'] != None:
         response = requests.get(tests['url'])
    else:
        response_body = json.loads(response.text)
        jsonpath_expression = parse(tests['Jpath'])
        for match in jsonpath_expression.find(response_body):
            assert match.value == tests['Expected']




