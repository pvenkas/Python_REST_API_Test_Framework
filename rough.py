import requests
import json
import openpyxl
from jsonpath_ng import jsonpath, parse


book = openpyxl.load_workbook("Attributes.xlsx")
sheet = book.active
testcases=[]
for row in range(2,sheet.max_row+1):
    test={}
    for column in range(1, sheet.max_column+1):
        #if sheet.cell(row=row,column=column).value != None:
        test[sheet.cell(row=1,column=column).value]=sheet.cell(row=row,column=column).value  # adding as dict
    testcases.append(test)
count = 0
for tests in testcases:
     if tests['url'] != None:
         response = requests.get(tests['url'])
         print("Get request")
     else:
         response_body = json.loads(response.text)
         jsonpath_expression = parse(tests['Jpath'])
         for match in jsonpath_expression.find(response_body):
             assert match.value == tests['Expected']

#print(testcases)



